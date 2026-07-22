import os
import io
import json
from datetime import date
from urllib.parse import quote
from fastapi import FastAPI, Request, Form, Depends, HTTPException, Query
from fastapi.responses import RedirectResponse, StreamingResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app import auth, repositorio, avaliacoes, repositorio_vinculo_tecnico

load_dotenv()

app = FastAPI(title="Avaliação de Técnicos")
app.add_middleware(SessionMiddleware, secret_key=os.getenv("SESSION_SECRET", "troque-esta-chave"))
app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")
templates.env.filters["normalizar"] = repositorio.normalizar
# Workaround: bug conhecido do Jinja2 no Python 3.14 quebra o cache interno
# de templates (https://github.com/pallets/jinja/issues/2180).
# Desabilitar o cache evita o erro "cannot use 'tuple' as a dict key".
templates.env.cache = None


@app.on_event("startup")
def _rodar_migracoes_no_startup():
    """Garante as colunas necessárias UMA VEZ, na subida do servidor —
    não a cada clique (ver o porquê em repositorio.rodar_migracoes_unicas)."""
    repositorio.rodar_migracoes_unicas()


def supervisor_logado(request: Request) -> str | None:
    return request.session.get("supervisor")


def _parse_data_opcional(valor: str | None) -> date | None:
    """Converte string de formulário em date, tratando "" (campo deixado em
    branco) como None. O FastAPI/Pydantic sozinho tenta parsear "" como data
    e quebra com erro 422 — por isso campos de data opcionais chegam aqui
    como str e passam por essa função antes de seguir pro banco."""
    if not valor or not valor.strip():
        return None
    return date.fromisoformat(valor.strip())


def exigir_login(request: Request):
    supervisor = supervisor_logado(request)
    if not supervisor:
        raise HTTPException(status_code=303, headers={"Location": "/login"})
    return supervisor


# ══════════════════════════════════════════════════════════
# LOGIN
# ══════════════════════════════════════════════════════════
@app.get("/login")
def tela_login(request: Request):
    if supervisor_logado(request):
        return RedirectResponse("/")
    return templates.TemplateResponse("login.html", {"request": request, "erro": None})


@app.post("/login")
def fazer_login(request: Request, login: str = Form(...), senha: str = Form(...)):
    usuario = auth.autenticar(login, senha)
    if usuario is None:
        return templates.TemplateResponse(
            "login.html", {"request": request, "erro": "Login ou senha inválidos."}
        )

    request.session["supervisor"] = usuario.supervisor
    request.session["login"] = usuario.login
    request.session["tipo"] = usuario.tipo

    if usuario.precisa_trocar_senha:
        return RedirectResponse("/trocar-senha", status_code=303)
    if usuario.tipo == "coordenador":
        return RedirectResponse("/coordenador", status_code=303)
    return RedirectResponse("/", status_code=303)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=303)


# ══════════════════════════════════════════════════════════
# TROCAR SENHA (obrigatório no primeiro acesso)
# ══════════════════════════════════════════════════════════
@app.get("/trocar-senha")
def tela_trocar_senha(request: Request):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    return templates.TemplateResponse("trocar_senha.html", {"request": request, "erro": None})


@app.post("/trocar-senha")
def trocar_senha(request: Request, nova_senha: str = Form(...), confirmar_senha: str = Form(...)):
    login = request.session.get("login")
    if not login:
        return RedirectResponse("/login", status_code=303)

    if nova_senha != confirmar_senha:
        return templates.TemplateResponse(
            "trocar_senha.html", {"request": request, "erro": "As senhas não conferem."}
        )
    if len(nova_senha) < 6:
        return templates.TemplateResponse(
            "trocar_senha.html", {"request": request, "erro": "A senha deve ter pelo menos 6 caracteres."}
        )

    auth.atualizar_senha(login, nova_senha)
    if request.session.get("tipo") == "coordenador":
        return RedirectResponse("/coordenador", status_code=303)
    return RedirectResponse("/", status_code=303)


def lista_meses_para_template():
    return [
        {"valor": m.strftime("%Y-%m"), "label": m.strftime("%m/%Y")}
        for m in repositorio.meses_disponiveis_para_avaliacao()
    ]


def _nova_planilha(cabecalho: list[str]):
    """Cria a planilha com o cabeçalho já formatado (negrito, fundo verde)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Avaliações"
    ws.append(cabecalho)

    fonte_cabecalho = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    fundo_cabecalho = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
    for celula in ws[1]:
        celula.font = fonte_cabecalho
        celula.fill = fundo_cabecalho
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    return wb, ws


def _ajustar_largura_colunas(ws, cabecalho: list[str]):
    for i, titulo in enumerate(cabecalho, start=1):
        largura = min(max(len(str(titulo)) + 4, 12), 45)
        ws.column_dimensions[get_column_letter(i)].width = largura


def _resposta_xlsx(wb: Workbook, nome_arquivo: str):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


def gerar_xlsx_response(linhas, nome_arquivo: str, incluir_supervisor: bool):
    """
    Monta a planilha Excel (.xlsx) simples: [supervisor,] técnico, média
    final e mês de referência.
    """
    cabecalho = ["Técnico", "Média Final", "Mês de Referência"]
    if incluir_supervisor:
        cabecalho.insert(0, "Supervisor")

    wb, ws = _nova_planilha(cabecalho)
    for r in linhas:
        media = float(r.nota_final) if r.nota_final is not None else None
        mes_label = r.mes_referencia.strftime("%m/%Y")
        linha = [r.tecnico, media, mes_label]
        if incluir_supervisor:
            linha.insert(0, r.supervisor)
        ws.append(linha)

    _ajustar_largura_colunas(ws, cabecalho)
    return _resposta_xlsx(wb, nome_arquivo)


def gerar_xlsx_completo_response(linhas: list[dict], nome_arquivo: str, incluir_supervisor: bool):
    """
    Monta a planilha Excel (.xlsx) completa: uma coluna para cada pergunta
    do formulário (rótulo curto quando existe, senão a pergunta por
    extenso), na ordem do formulário, além de supervisor/técnico/mês/média.
    """
    campos_ordenados = [
        (campo, avaliacoes.ROTULO_CURTO.get(campo, pergunta))
        for _titulo, campos in avaliacoes.PERGUNTAS
        for campo, pergunta in campos
    ]

    cabecalho = ["Técnico", "Mês de Referência"]
    if incluir_supervisor:
        cabecalho.insert(0, "Supervisor")
    cabecalho += [rotulo for _campo, rotulo in campos_ordenados]
    cabecalho.append("Média Final")

    wb, ws = _nova_planilha(cabecalho)
    for av in linhas:
        media = float(av["nota_final"]) if av.get("nota_final") is not None else None
        mes_label = av["mes_referencia"].strftime("%m/%Y")
        linha = [av["tecnico"], mes_label]
        if incluir_supervisor:
            linha.insert(0, av["supervisor"])
        for campo, _rotulo in campos_ordenados:
            linha.append(av.get(campo))
        linha.append(media)
        ws.append(linha)

    _ajustar_largura_colunas(ws, cabecalho)
    return _resposta_xlsx(wb, nome_arquivo)


# ══════════════════════════════════════════════════════════
# PAINEL DO SUPERVISOR
#   Passo 1: escolher o mês/ano de referência (tela "escolher_mes.html")
#   Passo 2: lista de técnicos daquele mês, com status avaliado/pendente
# ══════════════════════════════════════════════════════════
@app.get("/avaliacoes")
def painel(request: Request, mes: str | None = Query(default=None)):
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") == "coordenador":
        return RedirectResponse("/coordenador", status_code=303)

    meses_disponiveis = lista_meses_para_template()

    # Passo 1 — ainda não escolheu o mês: mostra a tela de seleção.
    if not mes:
        return templates.TemplateResponse(
            "escolher_mes.html",
            {
                "request": request,
                "supervisor": supervisor,
                "meses_disponiveis": meses_disponiveis,
                "destino": "/avaliacoes",
            },
        )

    # Passo 2 — mês escolhido: mostra os técnicos com o status daquele mês.
    mes_ref = resolver_mes_escolhido(mes)
    tecnicos = repositorio.tecnicos_com_status_para_mes(supervisor, mes_ref)
    ranking = repositorio.ranking_tecnicos_do_supervisor(supervisor, mes_ref)

    total = len(tecnicos)
    avaliados = sum(1 for t in tecnicos if t["avaliado"])
    pendentes = total - avaliados
    pct_conclusao = round(avaliados / total * 100, 1) if total else 0.0
    medias_validas = [r["nota_final"] for r in ranking if r["nota_final"] is not None]
    media_propria = round(sum(medias_validas) / len(medias_validas), 2) if medias_validas else None

    data_limite = repositorio.data_limite_do_mes(mes_ref)
    prazo_encerrado = date.today() > data_limite
    autorizado_apos_prazo = repositorio.existe_autorizacao(supervisor, mes_ref) if prazo_encerrado else False

    return templates.TemplateResponse(
        "painel.html",
        {
            "request": request,
            "supervisor": supervisor,
            "tecnicos": tecnicos,
            "ranking": ranking,
            "mes_referencia": mes_ref.strftime("%m/%Y"),
            "mes_valor": mes_ref.strftime("%Y-%m"),
            "total": total,
            "avaliados": avaliados,
            "pendentes": pendentes,
            "pct_conclusao": pct_conclusao,
            "media_propria": media_propria,
            "data_limite": data_limite.strftime("%d/%m/%Y"),
            "prazo_encerrado": prazo_encerrado,
            "autorizado_apos_prazo": autorizado_apos_prazo,
        },
    )


# ══════════════════════════════════════════════════════════
# DOWNLOAD — planilha simples (CSV) com técnico + média final
# ══════════════════════════════════════════════════════════
@app.get("/exportar")
def exportar_minhas_avaliacoes(request: Request, mes: str | None = Query(default=None)):
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)

    mes_ref = resolver_mes_escolhido(mes)
    linhas = repositorio.dados_para_exportar(mes_ref, supervisor=supervisor)
    nome_arquivo = f"minhas_avaliacoes_{mes_ref.strftime('%Y-%m')}.xlsx"
    return gerar_xlsx_response(linhas, nome_arquivo, incluir_supervisor=False)


@app.get("/exportar-completo")
def exportar_minhas_avaliacoes_completo(request: Request, mes: str | None = Query(default=None)):
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)

    mes_ref = resolver_mes_escolhido(mes)
    linhas = repositorio.avaliacoes_completas_mes(mes_ref, supervisor=supervisor)
    nome_arquivo = f"minhas_avaliacoes_completo_{mes_ref.strftime('%Y-%m')}.xlsx"
    return gerar_xlsx_completo_response(linhas, nome_arquivo, incluir_supervisor=False)


# ══════════════════════════════════════════════════════════
# HISTÓRICO DO SUPERVISOR — suas avaliações já feitas, por mês
# ══════════════════════════════════════════════════════════
@app.get("/minhas-avaliacoes")
def minhas_avaliacoes(request: Request):
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)

    eh_coordenador = request.session.get("tipo") == "coordenador"
    if eh_coordenador:
        grupos = repositorio.avaliacoes_geral()
    else:
        grupos = repositorio.avaliacoes_do_supervisor(supervisor)

    return templates.TemplateResponse(
        "historico.html",
        {
            "request": request,
            "supervisor": supervisor,
            "grupos": grupos,
            "eh_coordenador": eh_coordenador,
        },
    )


@app.get("/minhas-avaliacoes/{avaliacao_id}")
def detalhe_avaliacao(request: Request, avaliacao_id: int):
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)

    eh_coordenador = request.session.get("tipo") == "coordenador"
    if eh_coordenador:
        avaliacao = repositorio.buscar_avaliacao_por_id_admin(avaliacao_id)
    else:
        avaliacao = repositorio.buscar_avaliacao_por_id(avaliacao_id, supervisor)
    if avaliacao is None:
        raise HTTPException(status_code=404, detail="Avaliação não encontrada.")

    blocos = avaliacoes.montar_blocos_para_exibicao(avaliacao)
    resumo_visitas = repositorio.resumo_visitas_tecnico(avaliacao["tecnico"], avaliacao["mes_referencia"])
    return templates.TemplateResponse(
        "avaliacao_detalhe.html",
        {
            "request": request,
            "avaliacao": avaliacao,
            "blocos": blocos,
            "mes_referencia": avaliacao["mes_referencia"].strftime("%m/%Y"),
            "resumo_visitas": resumo_visitas,
        },
    )


# ══════════════════════════════════════════════════════════
# VISÃO DO COORDENADOR GERAL
#   Passo 1: escolher o mês/ano de referência (tela "escolher_mes.html")
#   Passo 2: lista de supervisores daquele mês, com totais de avaliados
#   Passo 3: ao clicar num supervisor, lista de técnicos + avaliações dele
# ══════════════════════════════════════════════════════════
@app.get("/coordenador")
def painel_coordenador_home(request: Request):
    """Tela inicial do coordenador: só os 4 links, sem pedir mês."""
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    return templates.TemplateResponse(
        "coordenador_home.html",
        {"request": request, "supervisor": request.session.get("supervisor")},
    )


@app.get("/coordenador/avaliacao")
def painel_coordenador(
    request: Request,
    mes: str | None = Query(default=None),
    ordenar: str | None = Query(default="nome"),
):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    meses_disponiveis = lista_meses_para_template()

    # Passo 1 — ainda não escolheu o mês: mostra a tela de seleção.
    if not mes:
        return templates.TemplateResponse(
            "escolher_mes.html",
            {
                "request": request,
                "supervisor": request.session.get("supervisor"),
                "meses_disponiveis": meses_disponiveis,
                "destino": "/coordenador/avaliacao",
            },
        )

    # Passo 2 — mês escolhido: mostra os supervisores com o status daquele mês.
    mes_ref = resolver_mes_escolhido(mes)
    resumo = repositorio.resumo_supervisores_para_mes(mes_ref)
    visao = repositorio.visao_geral_mes(mes_ref)

    ordenar = ordenar if ordenar in ("nome", "media", "pendentes") else "nome"
    if ordenar == "media":
        resumo = sorted(
            resumo,
            key=lambda s: (s["media_supervisor"] is None, -(s["media_supervisor"] or 0)),
        )
    elif ordenar == "pendentes":
        resumo = sorted(resumo, key=lambda s: (-s["pendentes"], s["supervisor"]))
    # "nome" já vem em ordem alfabética de resumo_supervisores_para_mes

    # Dados para o gráfico de barras (médias por supervisor), só entram
    # supervisores que já lançaram pelo menos uma avaliação no mês.
    # Convertido para float (em vez de Decimal) para desenhar a barra em CSS.
    maior_media = 10.0
    grafico_base = sorted(
        [
            {"supervisor": s["supervisor"], "media": float(s["media_supervisor"])}
            for s in resumo if s["media_supervisor"] is not None
        ],
        key=lambda s: s["media"],
        reverse=True,
    )
    for item in grafico_base:
        item["percentual"] = round(min(item["media"] / maior_media * 100, 100), 1)

    return templates.TemplateResponse(
        "coordenador.html",
        {
            "request": request,
            "resumo": resumo,
            "visao": visao,
            "ordenar": ordenar,
            "mes_referencia": mes_ref.strftime("%m/%Y"),
            "mes_valor": mes_ref.strftime("%Y-%m"),
            "grafico_base": grafico_base,
            "dia_limite": repositorio.obter_dia_limite(),
            "data_limite_mes": repositorio.data_limite_do_mes(mes_ref).strftime("%Y-%m-%d"),
            "data_limite_mes_exibicao": repositorio.data_limite_do_mes(mes_ref).strftime("%d/%m/%Y"),
            "solicitacoes_pendentes": repositorio.listar_solicitacoes_pendentes(),
        },
    )


# ══════════════════════════════════════════════════════════
# DOWNLOAD — planilha simples (CSV) com supervisor + técnico + média final
# Sem o parâmetro "supervisor": traz todos. Com ele: só daquele supervisor
# (usado no link da tela de técnicos de um supervisor específico).
# ══════════════════════════════════════════════════════════
@app.get("/coordenador/exportar")
def exportar_coordenador(
    request: Request,
    mes: str | None = Query(default=None),
    supervisor: str | None = Query(default=None),
):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    mes_ref = resolver_mes_escolhido(mes)
    linhas = repositorio.dados_para_exportar(mes_ref, supervisor=supervisor)

    if supervisor:
        nome_arquivo = f"avaliacoes_{supervisor}_{mes_ref.strftime('%Y-%m')}.xlsx"
    else:
        nome_arquivo = f"avaliacoes_geral_{mes_ref.strftime('%Y-%m')}.xlsx"

    return gerar_xlsx_response(linhas, nome_arquivo, incluir_supervisor=(not supervisor))


@app.get("/coordenador/exportar-completo")
def exportar_coordenador_completo(
    request: Request,
    mes: str | None = Query(default=None),
    supervisor: str | None = Query(default=None),
):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    mes_ref = resolver_mes_escolhido(mes)
    linhas = repositorio.avaliacoes_completas_mes(mes_ref, supervisor=supervisor)

    if supervisor:
        nome_arquivo = f"avaliacoes_completo_{supervisor}_{mes_ref.strftime('%Y-%m')}.xlsx"
    else:
        nome_arquivo = f"avaliacoes_completo_geral_{mes_ref.strftime('%Y-%m')}.xlsx"

    return gerar_xlsx_completo_response(linhas, nome_arquivo, incluir_supervisor=(not supervisor))


# ══════════════════════════════════════════════════════════
# COORDENADOR — ranking geral do mês (supervisores e técnicos)
# IMPORTANTE: esta rota tem que vir ANTES de /coordenador/{supervisor},
# senão "ranking" seria interpretado como nome de supervisor.
# ══════════════════════════════════════════════════════════
@app.get("/")
def tela_gestao_tecnicos(request: Request):
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") == "coordenador":
        return RedirectResponse("/coordenador", status_code=303)

    # Auto-vincular/auto-desvincular DESATIVADOS (2026-07): vínculo e
    # desvínculo técnico-supervisor agora são sempre feitos manualmente
    # pelo coordenador, em /coordenador/cadastros. O supervisor só
    # visualiza a equipe aqui — sem nenhuma ação de vincular/desvincular
    # no próprio perfil. O histórico de desvínculos fica em página
    # separada (/tecnicos-desvinculados) pra não poluir esta tela.

    todos_vinculos = repositorio_vinculo_tecnico.listar_vinculos_do_supervisor(supervisor)
    return templates.TemplateResponse(
        "gestao_tecnicos.html",
        {
            "request": request,
            "supervisor": supervisor,
            "vinculados": [v for v in todos_vinculos if v["data_desvinculacao"] is None],
            "total_desvinculados": sum(1 for v in todos_vinculos if v["data_desvinculacao"] is not None),
        },
    )


@app.get("/tecnicos-desvinculados")
def tela_tecnicos_desvinculados(request: Request):
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") == "coordenador":
        return RedirectResponse("/coordenador", status_code=303)

    todos_vinculos = repositorio_vinculo_tecnico.listar_vinculos_do_supervisor(supervisor)
    desvinculados = sorted(
        (v for v in todos_vinculos if v["data_desvinculacao"] is not None),
        key=lambda v: v["data_desvinculacao"],
        reverse=True,
    )
    return templates.TemplateResponse(
        "tecnicos_desvinculados.html",
        {
            "request": request,
            "supervisor": supervisor,
            "desvinculados": desvinculados,
        },
    )


@app.get("/tecnicos/{tecnico}")
def tela_detalhe_tecnico(request: Request, tecnico: str):
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)

    # Resolve o nome EXATO do técnico (o que vem da URL pode ter acentuação
    # ou espaçamento levemente diferente do que está gravado no banco),
    # pra garantir que a busca de visita/vínculo funcione certinho.
    tecnico_real = repositorio_vinculo_tecnico.resolver_nome_tecnico(tecnico)
    if tecnico_real is not None:
        tecnico = tecnico_real

    vinculo_ativo = repositorio_vinculo_tecnico.obter_vinculo_ativo_do_tecnico(tecnico)
    visita = repositorio_vinculo_tecnico.obter_visita_tecnico(tecnico)
    historico = repositorio_vinculo_tecnico.historico_tecnico(tecnico)
    tecnico_cadastro = repositorio.obter_tecnico_por_nome(tecnico)

    eh_meu = vinculo_ativo is not None and vinculo_ativo["supervisor"] == supervisor
    eh_coordenador = request.session.get("tipo") == "coordenador"
    return templates.TemplateResponse(
        "detalhe_tecnico.html",
        {
            "request": request,
            "supervisor": supervisor,
            "tecnico": tecnico,
            "vinculo_ativo": vinculo_ativo,
            "eh_meu": eh_meu,
            "eh_coordenador": eh_coordenador,
            "todos_supervisores": repositorio.listar_supervisores() if eh_coordenador else [],
            "visita": visita,
            "historico": [h for h in historico if h["data_desvinculacao"] is not None],
            "motivos": repositorio_vinculo_tecnico.MOTIVOS_DESVINCULACAO,
            "tecnico_cadastro": tecnico_cadastro,
            "motivos_desativacao_tecnico": repositorio.MOTIVOS_DESATIVACAO_TECNICO,
        },
    )


@app.post("/tecnicos/criar")
def criar_vinculo_tecnico(
    request: Request,
    tecnico: str = Form(...),
    cpf: str = Form(""),
    projeto: str = Form(""),
    atividade: str = Form(""),
    empresa: str = Form(""),
    cnpj_empresa: str = Form(""),
    data_inicio: date = Form(...),
    data_fim_prevista: str | None = Form(None),
    supervisor_escolhido: str | None = Form(default=None),
):
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)

    # Só o coordenador credencia (vincula) técnico — supervisor não decide
    # mais isso sozinho, só o coordenador em /coordenador/cadastros/associar
    # (ou aqui mesmo, escolhendo o supervisor no formulário).
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    supervisor_do_vinculo = supervisor_escolhido or supervisor

    repositorio_vinculo_tecnico.criar_vinculo(
        tecnico=tecnico,
        supervisor=supervisor_do_vinculo,
        data_inicio=data_inicio,
        data_fim_prevista=_parse_data_opcional(data_fim_prevista),
        cpf=cpf,
        projeto=projeto,
        atividade=atividade,
        empresa=empresa,
        cnpj_empresa=cnpj_empresa,
        criado_por=request.session.get("login", supervisor),
    )
    return RedirectResponse(f"/tecnicos/{tecnico}", status_code=303)


@app.post("/tecnicos/editar")
def editar_vinculo_tecnico(
    request: Request,
    vinculo_id: int = Form(...),
    tecnico: str = Form(...),
    data_inicio: str | None = Form(None),
    data_fim_prevista: str | None = Form(None),
):
    """Edita só o início e o fim previsto do vínculo ativo — projeto e
    atividade vêm da base de visitas e não são editáveis aqui."""
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)

    repositorio_vinculo_tecnico.editar_datas_vinculo(
        vinculo_id=vinculo_id,
        data_inicio=_parse_data_opcional(data_inicio),
        data_fim_prevista=_parse_data_opcional(data_fim_prevista),
    )
    return RedirectResponse(f"/tecnicos/{tecnico}", status_code=303)


@app.post("/tecnicos/mudar-supervisor")
def mudar_supervisor_tecnico(
    request: Request,
    vinculo_id: int = Form(...),
    tecnico: str = Form(...),
    novo_supervisor: str = Form(...),
    data_inicio: date = Form(...),
    data_fim_prevista: str | None = Form(None),
):
    """Só o coordenador vê esse botão — troca o supervisor de um técnico
    em um único passo, mantendo projeto/atividade/empresa/CPF do vínculo atual.
    O motivo é sempre "Mudança de supervisor": é a única razão de existir
    dessa ação (pra outros motivos de desvínculo, use "Desativar Técnico")."""
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    vinculo_atual = repositorio_vinculo_tecnico.obter_vinculo(vinculo_id)
    if vinculo_atual is None:
        raise HTTPException(status_code=404, detail="Vínculo não encontrado.")

    repositorio_vinculo_tecnico.desvincular_vinculo(
        vinculo_id=vinculo_id,
        data_desvinculacao=data_inicio,
        motivo="Mudança de supervisor",
    )
    repositorio_vinculo_tecnico.criar_vinculo(
        tecnico=tecnico,
        supervisor=novo_supervisor,
        data_inicio=data_inicio,
        criado_por=request.session.get("login", "coordenador"),
        cpf=vinculo_atual.get("cpf"),
        projeto=vinculo_atual.get("projeto"),
        atividade=vinculo_atual.get("atividade"),
        empresa=vinculo_atual.get("empresa"),
        cnpj_empresa=vinculo_atual.get("cnpj_empresa"),
        data_fim_prevista=_parse_data_opcional(data_fim_prevista),
    )
    return RedirectResponse(f"/tecnicos/{tecnico}", status_code=303)


@app.post("/tecnicos/{tecnico}/desativar")
def desativar_tecnico_na_pagina(
    request: Request,
    tecnico: str,
    id_tecnico: int = Form(...),
    motivo_desativacao: str = Form(...),
):
    """
    Desativa o técnico (tabela mestra 'tecnicos') direto da página de
    detalhe dele, com motivo obrigatório. Se ele tiver um vínculo ativo,
    esse vínculo também é encerrado automaticamente — um técnico
    desativado não fica com vínculo aberto.
    """
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    tecnico_real = repositorio_vinculo_tecnico.resolver_nome_tecnico(tecnico)
    if tecnico_real is not None:
        tecnico = tecnico_real

    repositorio.definir_ativo_tecnico(id_tecnico, False, motivo_desativacao)

    vinculo_ativo = repositorio_vinculo_tecnico.obter_vinculo_ativo_do_tecnico(tecnico)
    if vinculo_ativo is not None:
        repositorio_vinculo_tecnico.desvincular_vinculo(
            vinculo_id=vinculo_ativo["id"],
            data_desvinculacao=date.today(),
            motivo=f"Técnico desativado: {motivo_desativacao}",
        )

    return RedirectResponse(f"/tecnicos/{tecnico}", status_code=303)


@app.post("/tecnicos/desvincular")
def desvincular_vinculo_tecnico(
    request: Request,
    vinculo_id: int = Form(...),
    tecnico: str = Form(...),
    motivo_desvinculacao: str = Form(...),
    data_desvinculacao: date = Form(...),
    redirecionar_para: str | None = Form(default=None),
):
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)

    # Só o coordenador descredencia (desvincula) técnico — igual ao
    # credenciamento, supervisor não decide isso mais sozinho.
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    repositorio_vinculo_tecnico.desvincular_vinculo(
        vinculo_id=vinculo_id, data_desvinculacao=data_desvinculacao, motivo=motivo_desvinculacao
    )
    # Volta pra onde a ação foi disparada: da lista de técnicos, volta pra lista
    # (onde o técnico já aparece em "Não vinculado"); do detalhe do técnico,
    # continua no detalhe (que já mostra o formulário de complementar cadastro).
    destino = redirecionar_para or f"/tecnicos/{tecnico}"
    return RedirectResponse(destino, status_code=303)


@app.get("/coordenador/tecnicos-desativados")
def tela_tecnicos_desativados_coordenador(request: Request):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    return templates.TemplateResponse(
        "tecnicos_desativados_coordenador.html",
        {
            "request": request,
            "supervisor": request.session.get("supervisor"),
            "desativados": repositorio.listar_tecnicos_desativados(),
            "historico_desvinculacoes": repositorio_vinculo_tecnico.listar_historico_desvinculacoes(),
        },
    )


@app.get("/coordenador/tecnicos/relatorio")
def tela_relatorio_completo_tecnicos(request: Request, supervisor: str | None = Query(default=None)):
    """Relatório com TODO técnico cadastrado, dados cadastrais + situação
    atual (ativo/desativado/descredenciado) — pra visualizar e imprimir."""
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    return templates.TemplateResponse(
        "relatorio_tecnicos.html",
        {
            "request": request,
            "supervisor": request.session.get("supervisor"),
            "supervisores": repositorio.listar_supervisores_cadastrados(apenas_ativos=True),
            "supervisor_escolhido": supervisor,
            "tecnicos": repositorio.listar_relatorio_completo_tecnicos(supervisor),
        },
    )


@app.get("/coordenador/tecnicos/relatorio.xlsx")
def baixar_relatorio_completo_tecnicos(request: Request, supervisor: str | None = Query(default=None)):
    """Mesmo relatório de cima, em Excel pra baixar."""
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    cabecalho = [
        "Nome", "Situação", "Projeto", "Atividade", "Motivo (desativação)", "Avaliação do Técnico",
        "RG", "CPF", "Contato", "E-mail",
        "Endereço", "Município", "Empresa", "CNPJ da empresa",
        "Capacitação metodológica (mês/ano)", "Modalidade da capacitação",
    ]
    wb, ws = _nova_planilha(cabecalho)
    ws.title = "Técnicos"
    for t in repositorio.listar_relatorio_completo_tecnicos(supervisor):
        capacitacao = t.get("mes_ano_capacitacao_metodologica")
        ws.append([
            t["nome"], t["situacao"], t.get("projeto_atual") or "", t.get("atividade_atual") or "",
            t.get("motivo_desativacao_curto") or "",
            t.get("avaliacao_desativacao") or "",
            t.get("rg") or "", t.get("cpf") or "",
            t.get("contato") or "", t.get("email") or "", t.get("endereco") or "",
            t.get("municipio") or "", t.get("empresa") or "", t.get("cnpj_empresa") or "",
            capacitacao.strftime("%m/%Y") if capacitacao else "",
            t.get("modalidade_capacitacao_metodologica") or "",
        ])
    _ajustar_largura_colunas(ws, cabecalho)
    nome_arquivo = f"relatorio_tecnicos_{supervisor}.xlsx" if supervisor else "relatorio_tecnicos.xlsx"
    return _resposta_xlsx(wb, nome_arquivo)


@app.post("/coordenador/tecnicos-desativados/{id_tecnico}/reativar")
def reativar_tecnico_coordenador(request: Request, id_tecnico: int):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    repositorio.definir_ativo_tecnico(id_tecnico, True)
    return RedirectResponse("/coordenador/tecnicos-desativados", status_code=303)


@app.get("/coordenador/vinculos-tecnicos")
def tela_vinculos_tecnicos_coordenador(request: Request):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    vinculos = repositorio_vinculo_tecnico.listar_todos_vinculos_ativos()
    return templates.TemplateResponse(
        "vinculos_tecnicos_coordenador.html",
        {
            "request": request,
            "supervisor": request.session.get("supervisor"),
            "vinculos": vinculos,
        },
    )


@app.get("/coordenador/cadastros/associar")
def tela_associar_tecnicos(request: Request, supervisor: str | None = Query(default=None)):
    """Escolhe um supervisor, marca os técnicos dele (dos cadastros mestres) e associa em lote."""
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    supervisores = repositorio.listar_supervisores_cadastrados(apenas_ativos=True)
    todos_tecnicos = repositorio.listar_tecnicos_cadastrados(apenas_ativos=True)

    # Vínculo ativo com QUALQUER supervisor (não só o selecionado) — um
    # técnico já vinculado a outro supervisor não pode aparecer como
    # "Disponível" aqui, senão o coordenador marca ele achando que está
    # livre e a associação falha silenciosamente (índice único do banco).
    vinculos_ativos = repositorio_vinculo_tecnico.listar_todos_vinculos_ativos()
    supervisor_atual_por_tecnico = {
        repositorio.normalizar(v["tecnico"]): v["supervisor"] for v in vinculos_ativos
    }

    return templates.TemplateResponse(
        "associar_tecnicos.html",
        {
            "request": request,
            "supervisores": supervisores,
            "tecnicos": todos_tecnicos,
            "supervisor_escolhido": supervisor,
            "supervisor_atual_por_tecnico": supervisor_atual_por_tecnico,
        },
    )


@app.post("/coordenador/cadastros/associar")
async def associar_tecnicos_em_lote(request: Request):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    form = await request.form()
    supervisor = form.get("supervisor")
    tecnicos_ids = form.getlist("tecnicos")
    coordenador_login = request.session.get("login", "coordenador")

    associados = 0
    for id_tecnico in tecnicos_ids:
        tecnico_cadastro = repositorio.obter_tecnico(int(id_tecnico))
        if not tecnico_cadastro:
            continue
        try:
            repositorio_vinculo_tecnico.criar_vinculo(
                tecnico=tecnico_cadastro["nome"],
                supervisor=supervisor,
                data_inicio=date.today(),
                criado_por=coordenador_login,
                cpf=tecnico_cadastro.get("cpf"),
                empresa=tecnico_cadastro.get("empresa"),
                cnpj_empresa=tecnico_cadastro.get("cnpj_empresa"),
            )
            associados += 1
        except Exception:
            # Provavelmente já está ativo com esse ou outro supervisor — pula.
            continue

    return RedirectResponse(f"/coordenador/cadastros/associar?supervisor={supervisor}", status_code=303)


@app.post("/coordenador/cadastros/tecnico/descredenciar")
def descredenciar_tecnico_coordenador(
    request: Request,
    vinculo_id: int = Form(...),
    motivo_desvinculacao: str = Form(...),
    novo_supervisor: str = Form(""),
    data_inicio_novo: str | None = Form(None),
    data_fim_prevista_novo: str | None = Form(None),
):
    coordenador_login = supervisor_logado(request)
    if not coordenador_login:
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    # Pega os dados do vínculo ANTES de encerrar, pra poder recriar no
    # supervisor novo (mesmo técnico, cpf, empresa, cnpj) se for transferência.
    vinculo_antigo = repositorio_vinculo_tecnico.obter_vinculo(vinculo_id)

    repositorio_vinculo_tecnico.desvincular_vinculo(
        vinculo_id=vinculo_id, data_desvinculacao=date.today(), motivo=motivo_desvinculacao
    )

    if novo_supervisor and vinculo_antigo:
        repositorio_vinculo_tecnico.criar_vinculo(
            tecnico=vinculo_antigo["tecnico"],
            supervisor=novo_supervisor,
            data_inicio=_parse_data_opcional(data_inicio_novo) or date.today(),
            data_fim_prevista=_parse_data_opcional(data_fim_prevista_novo),
            criado_por=coordenador_login,
            cpf=vinculo_antigo.get("cpf"),
            empresa=vinculo_antigo.get("empresa"),
            cnpj_empresa=vinculo_antigo.get("cnpj_empresa"),
        )

    return RedirectResponse("/coordenador/cadastros", status_code=303)


@app.post("/coordenador/cadastros/tecnico/reverter")
def reverter_descredenciamento_coordenador(request: Request, vinculo_id: int = Form(...)):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    try:
        repositorio_vinculo_tecnico.reverter_desvinculacao(vinculo_id)
    except Exception:
        # Provavelmente o técnico já tem outro vínculo ativo (foi
        # transferido depois) — não dá pra reverter sem tirar de lá primeiro.
        # Busca com qual supervisor ele está agora pra deixar isso claro
        # em vez de só dizer "outro vínculo ativo" sem dizer qual.
        mensagem = "Não foi possível reverter: esse técnico já tem outro vínculo ativo. Descredencie o vínculo atual dele primeiro."
        vinculo_antigo = repositorio_vinculo_tecnico.obter_vinculo(vinculo_id)
        if vinculo_antigo:
            vinculo_atual = repositorio_vinculo_tecnico.obter_vinculo_ativo_do_tecnico(vinculo_antigo["tecnico"])
            if vinculo_atual:
                mensagem = (
                    f"Não foi possível reverter: {vinculo_antigo['tecnico']} já está vinculado a "
                    f"{vinculo_atual['supervisor']} desde "
                    f"{vinculo_atual['data_inicio'].strftime('%d/%m/%Y') if vinculo_atual.get('data_inicio') else '—'}. "
                    f"Descredencie esse vínculo atual primeiro."
                )
        return RedirectResponse(
            f"/coordenador/cadastros?erro={quote(mensagem)}",
            status_code=303,
        )
    return RedirectResponse("/coordenador/cadastros", status_code=303)


@app.get("/coordenador/cadastros")
def tela_cadastros(request: Request):
    """Tela do coordenador: inserir supervisor manualmente e ver/ativar/
    desativar supervisores. (A parte de técnicos fica de fora por enquanto,
    só a de supervisores está em uso nesta aba.)"""
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    # Cronômetros temporários (2026-07) só pra investigar a lentidão dessa
    # tela em produção — remover depois que o gargalo for identificado.
    import time as _time
    _t0 = _time.time()

    # Traz pra tabela 'supervisores' quem já tem login como tipo 'supervisor'
    # em usuarios_supervisores (não duplica quem já estava aqui).
    repositorio.sincronizar_supervisores_de_usuarios()
    _t1 = _time.time()
    print(f"[TIMING /coordenador/cadastros] sincronizar_supervisores_de_usuarios: {_t1 - _t0:.2f}s", flush=True)

    supervisores = repositorio.listar_supervisores_cadastrados()
    _t2 = _time.time()
    print(f"[TIMING /coordenador/cadastros] listar_supervisores_cadastrados: {_t2 - _t1:.2f}s", flush=True)

    vinculos_agrupados = repositorio_vinculo_tecnico.listar_vinculos_de_todos_os_supervisores()
    _t3 = _time.time()
    print(f"[TIMING /coordenador/cadastros] listar_vinculos_de_todos_os_supervisores: {_t3 - _t2:.2f}s", flush=True)

    vinculos_por_supervisor = {
        s["id"]: vinculos_agrupados.get(s["nome"], [])
        for s in supervisores
    }
    print(f"[TIMING /coordenador/cadastros] TOTAL ate aqui: {_t3 - _t0:.2f}s", flush=True)
    return templates.TemplateResponse(
        "cadastros_coordenador.html",
        {
            "request": request,
            "supervisores": supervisores,
            "motivos_desativacao_supervisor": repositorio.MOTIVOS_DESATIVACAO_SUPERVISOR,
            "motivos_desvinculacao_tecnico": repositorio_vinculo_tecnico.MOTIVOS_DESVINCULACAO,
            "motivos_desativacao_tecnico": repositorio.MOTIVOS_DESATIVACAO_TECNICO,
            "projetos_por_supervisor": {
                s["id"]: repositorio.listar_projetos_do_supervisor(s["nome"])
                for s in supervisores
            },
            "tecnicos_vinculados_por_supervisor": {
                s["id"]: repositorio.contar_tecnicos_vinculados(s["nome"])
                for s in supervisores
            },
            "vinculos_ativos_por_supervisor": {
                sid: [v for v in vinculos if v["data_desvinculacao"] is None]
                for sid, vinculos in vinculos_por_supervisor.items()
            },
            "historico_por_supervisor": {
                sid: [v for v in vinculos if v["data_desvinculacao"] is not None]
                for sid, vinculos in vinculos_por_supervisor.items()
            },
            "historico_desativados_por_supervisor": {
                sid: [
                    v for v in vinculos
                    if v["data_desvinculacao"] is not None
                    and (v["motivo_desvinculacao"] or "").startswith("Técnico desativado")
                ]
                for sid, vinculos in vinculos_por_supervisor.items()
            },
            "historico_descredenciados_por_supervisor": {
                sid: [
                    v for v in vinculos
                    if v["data_desvinculacao"] is not None
                    and not (v["motivo_desvinculacao"] or "").startswith("Técnico desativado")
                ]
                for sid, vinculos in vinculos_por_supervisor.items()
            },
            "supervisores_ativos_nomes": [s["nome"] for s in supervisores if s["ativo"]],
            "erro": request.query_params.get("erro"),
        },
    )


@app.post("/coordenador/cadastros/supervisor/criar")
def criar_supervisor_cadastro(
    request: Request,
    nome: str = Form(...),
    rg: str = Form(""),
    cpf: str = Form(""),
    contato: str = Form(""),
    empresa: str = Form(""),
    cnpj_empresa: str = Form(""),
    data_inicio_vinculo: str | None = Form(None),
    data_fim_vinculo: str | None = Form(None),
):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    repositorio.criar_supervisor(
        nome, rg=rg, cpf=cpf, contato=contato, empresa=empresa, cnpj_empresa=cnpj_empresa,
        data_inicio_vinculo=_parse_data_opcional(data_inicio_vinculo),
        data_fim_vinculo=_parse_data_opcional(data_fim_vinculo),
    )
    return RedirectResponse("/coordenador/cadastros", status_code=303)


@app.post("/coordenador/cadastros/supervisor/{supervisor_id}/excluir")
def excluir_supervisor_cadastro(request: Request, supervisor_id: int, nome: str = Form(...)):
    """Exclui de vez o cadastro do supervisor — só permitido se ele não
    tiver nenhum técnico vinculado no momento."""
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    try:
        repositorio.excluir_supervisor(supervisor_id, nome)
    except ValueError:
        raise HTTPException(status_code=400, detail="Não é possível excluir: supervisor ainda tem técnico(s) vinculado(s).")
    return RedirectResponse("/coordenador/cadastros", status_code=303)


@app.post("/coordenador/cadastros/supervisor/{supervisor_id}/ativo")
def alterar_ativo_supervisor(
    request: Request,
    supervisor_id: int,
    ativo: bool = Form(...),
    motivo_desativacao: str = Form(""),
):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    repositorio.definir_ativo_supervisor(supervisor_id, ativo, motivo_desativacao)
    return RedirectResponse("/coordenador/cadastros", status_code=303)


@app.post("/coordenador/cadastros/tecnico/{id_tecnico}/ativo")
def alterar_ativo_tecnico(request: Request, id_tecnico: int, ativo: bool = Form(...), motivo_desativacao: str = Form("")):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    repositorio.definir_ativo_tecnico(id_tecnico, ativo, motivo_desativacao)

    # Um técnico desativado não fica com vínculo aberto — encerra
    # automaticamente e ele passa a aparecer no Histórico, não mais na Equipe.
    if not ativo:
        tecnico_cadastro = repositorio.obter_tecnico(id_tecnico)
        if tecnico_cadastro is not None:
            vinculo_ativo = repositorio_vinculo_tecnico.obter_vinculo_ativo_do_tecnico(tecnico_cadastro["nome"])
            if vinculo_ativo is not None:
                repositorio_vinculo_tecnico.desvincular_vinculo(
                    vinculo_id=vinculo_ativo["id"],
                    data_desvinculacao=date.today(),
                    motivo=f"Técnico desativado: {motivo_desativacao}" if motivo_desativacao else "Técnico desativado",
                )

    return RedirectResponse("/coordenador/cadastros", status_code=303)


@app.post("/coordenador/cadastros/tecnico/{id_tecnico}/editar")
def editar_dados_cadastrais_tecnico(
    request: Request,
    id_tecnico: int,
    rg: str = Form(""),
    cpf: str = Form(""),
    contato: str = Form(""),
    email: str = Form(""),
    empresa: str = Form(""),
    cnpj_empresa: str = Form(""),
    endereco: str = Form(""),
    municipio: str = Form(""),
    data_inicio_vinculo: str | None = Form(None),
    data_fim_vinculo: str | None = Form(None),
    mes_ano_capacitacao_metodologica: str | None = Form(None),
    modalidade_capacitacao_metodologica: str = Form(""),
):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    mes_ano_capacitacao = None
    if mes_ano_capacitacao_metodologica:
        mes_ano_capacitacao = _parse_data_opcional(mes_ano_capacitacao_metodologica + "-01")

    repositorio.atualizar_dados_cadastrais_tecnico(
        id_tecnico_responsavel=id_tecnico,
        rg=rg, cpf=cpf, contato=contato, email=email, empresa=empresa, cnpj_empresa=cnpj_empresa,
        endereco=endereco, municipio=municipio,
        data_inicio_vinculo=_parse_data_opcional(data_inicio_vinculo),
        data_fim_vinculo=_parse_data_opcional(data_fim_vinculo),
        mes_ano_capacitacao_metodologica=mes_ano_capacitacao,
        modalidade_capacitacao_metodologica=modalidade_capacitacao_metodologica or None,
    )
    return RedirectResponse("/coordenador/cadastros", status_code=303)


@app.post("/tecnicos/{tecnico}/cadastro/editar")
def editar_dados_cadastrais_tecnico_na_pagina(
    request: Request,
    tecnico: str,
    id_tecnico: int = Form(...),
    rg: str = Form(""),
    cpf: str = Form(""),
    contato: str = Form(""),
    email: str = Form(""),
    empresa: str = Form(""),
    cnpj_empresa: str = Form(""),
    endereco: str = Form(""),
    municipio: str = Form(""),
    data_inicio_vinculo: str | None = Form(None),
    data_fim_vinculo: str | None = Form(None),
    mes_ano_capacitacao_metodologica: str | None = Form(None),
    modalidade_capacitacao_metodologica: str = Form(...),
):
    """
    Mesma atualização de dados cadastrais do técnico (tabela mestra
    'tecnicos'), só que disparada direto da página /tecnicos/{tecnico}
    (em vez da tela de gestão do coordenador), voltando pra essa mesma
    página depois de salvar.
    """
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)

    tecnico_real = repositorio_vinculo_tecnico.resolver_nome_tecnico(tecnico)
    if tecnico_real is not None:
        tecnico = tecnico_real

    # Só quem é responsável pelo técnico agora ou o coordenador geral
    # pode editar os dados cadastrais dele.
    vinculo_ativo = repositorio_vinculo_tecnico.obter_vinculo_ativo_do_tecnico(tecnico)
    eh_meu = vinculo_ativo is not None and vinculo_ativo["supervisor"] == supervisor
    eh_coordenador = request.session.get("tipo") == "coordenador"
    if not (eh_meu or eh_coordenador):
        raise HTTPException(status_code=403, detail="Sem permissão pra editar o cadastro deste técnico.")

    # O input type="month" do formulário manda "AAAA-MM" — completa com o
    # dia 01 pra virar uma data válida (guardamos só mês/ano mesmo).
    mes_ano_capacitacao = None
    if mes_ano_capacitacao_metodologica:
        mes_ano_capacitacao = _parse_data_opcional(mes_ano_capacitacao_metodologica + "-01")

    repositorio.atualizar_dados_cadastrais_tecnico(
        id_tecnico_responsavel=id_tecnico,
        rg=rg, cpf=cpf, contato=contato, email=email, empresa=empresa, cnpj_empresa=cnpj_empresa,
        endereco=endereco, municipio=municipio,
        data_inicio_vinculo=_parse_data_opcional(data_inicio_vinculo),
        data_fim_vinculo=_parse_data_opcional(data_fim_vinculo),
        mes_ano_capacitacao_metodologica=mes_ano_capacitacao,
        modalidade_capacitacao_metodologica=modalidade_capacitacao_metodologica,
    )
    return RedirectResponse(f"/tecnicos/{tecnico}", status_code=303)


@app.post("/coordenador/cadastros/supervisor/{supervisor_id}/editar")
def editar_dados_cadastrais_supervisor(
    request: Request,
    supervisor_id: int,
    rg: str = Form(""),
    cpf: str = Form(""),
    contato: str = Form(""),
    empresa: str = Form(""),
    cnpj_empresa: str = Form(""),
    data_inicio_vinculo: str | None = Form(None),
    data_fim_vinculo: str | None = Form(None),
):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    repositorio.atualizar_dados_cadastrais_supervisor(
        supervisor_id=supervisor_id,
        rg=rg, cpf=cpf, contato=contato, empresa=empresa, cnpj_empresa=cnpj_empresa,
        data_inicio_vinculo=_parse_data_opcional(data_inicio_vinculo),
        data_fim_vinculo=_parse_data_opcional(data_fim_vinculo),
    )
    return RedirectResponse("/coordenador/cadastros", status_code=303)


@app.post("/coordenador/cadastros/sincronizar")
def sincronizar_tecnicos_visitas(request: Request, redirecionar_para: str | None = Form(default=None)):
    """Botão 'Atualizar da base de visitas' — traz técnicos novos/atualiza
    nome e datas (todo o histórico até hoje, sem precisar escolher período),
    e já sincroniza as combinações projeto+atividade de cada um."""
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    repositorio.sincronizar_tecnicos_da_visita()
    repositorio.sincronizar_tecnico_atividades()
    return RedirectResponse(redirecionar_para or "/coordenador/cadastros", status_code=303)


@app.get("/coordenador/ranking")
def ranking_coordenador(request: Request, mes: str | None = Query(default=None)):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    mes_ref = resolver_mes_escolhido(mes)
    ranking_supervisores = repositorio.ranking_supervisores(mes_ref)
    ranking_tecnicos = repositorio.ranking_geral_tecnicos(mes_ref)

    return templates.TemplateResponse(
        "ranking_coordenador.html",
        {
            "request": request,
            "ranking_supervisores": ranking_supervisores,
            "ranking_tecnicos": ranking_tecnicos,
            "mes_referencia": mes_ref.strftime("%m/%Y"),
            "mes_valor": mes_ref.strftime("%Y-%m"),
        },
    )


@app.get("/coordenador/ranking-objetivo")
def ranking_objetivo_coordenador(request: Request, mes: str | None = Query(default=None)):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    mes_ref = resolver_mes_escolhido(mes)

    # Ranking objetivo (visitas) — MESMO mês escolhido pra avaliação subjetiva
    # (ex: escolheu junho -> só considera visitas de 01/06 a 30/06).
    import calendar
    ultimo_dia = calendar.monthrange(mes_ref.year, mes_ref.month)[1]
    obj_dt_inicio = mes_ref
    obj_dt_fim = date(mes_ref.year, mes_ref.month, ultimo_dia)

    ranking_objetivo_rows = repositorio.ranking_objetivo_tecnicos(obj_dt_inicio, obj_dt_fim)

    # Agrupa por supervisor, já ordenado pela posição (pos) dentro do grupo.
    ranking_objetivo_por_supervisor = {}
    for r in ranking_objetivo_rows:
        chave = r["ultimo_supervisor"] or "— sem supervisor —"
        ranking_objetivo_por_supervisor.setdefault(chave, []).append(r)
    for lista in ranking_objetivo_por_supervisor.values():
        lista.sort(key=lambda r: r["pos"])
    ranking_objetivo_por_supervisor = dict(sorted(ranking_objetivo_por_supervisor.items()))

    return templates.TemplateResponse(
        "ranking_objetivo.html",
        {
            "request": request,
            "mes_referencia": mes_ref.strftime("%m/%Y"),
            "mes_valor": mes_ref.strftime("%Y-%m"),
            "ranking_objetivo_por_supervisor": ranking_objetivo_por_supervisor,
        },
    )


# ══════════════════════════════════════════════════════════
# COORDENADOR — técnicos e avaliações de UM supervisor, no mês escolhido
# ══════════════════════════════════════════════════════════
@app.get("/coordenador/{supervisor}")
def painel_coordenador_supervisor(request: Request, supervisor: str, mes: str | None = Query(default=None)):
    if not supervisor_logado(request):
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    mes_ref = resolver_mes_escolhido(mes)
    tecnicos = repositorio.tecnicos_com_avaliacao_para_mes(supervisor, mes_ref)

    prazo_encerrado = repositorio.prazo_do_mes_encerrado(mes_ref)
    autorizado = repositorio.existe_autorizacao(supervisor, mes_ref) if prazo_encerrado else False

    return templates.TemplateResponse(
        "coordenador_supervisor.html",
        {
            "request": request,
            "supervisor": supervisor,
            "tecnicos": tecnicos,
            "mes_referencia": mes_ref.strftime("%m/%Y"),
            "mes_valor": mes_ref.strftime("%Y-%m"),
            "prazo_encerrado": prazo_encerrado,
            "autorizado": autorizado,
            "data_limite": repositorio.data_limite_do_mes(mes_ref).strftime("%d/%m/%Y"),
        },
    )


# ══════════════════════════════════════════════════════════
# PRAZO DE AVALIAÇÃO — configuração do dia-limite (coordenador)
# e autorização pontual de um supervisor fora do prazo
# ══════════════════════════════════════════════════════════
@app.post("/coordenador/configurar-prazo")
def configurar_prazo(request: Request, dia_limite: int = Form(...), mes: str | None = Form(default=None)):
    coordenador = supervisor_logado(request)
    if not coordenador:
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    dia_limite = max(1, min(31, dia_limite))  # protege contra valor absurdo vindo do form
    repositorio.definir_dia_limite(dia_limite, coordenador)

    destino = "/coordenador" + (f"?mes={mes}" if mes else "")
    return RedirectResponse(destino, status_code=303)


@app.post("/coordenador/definir-prazo-do-mes")
def definir_prazo_do_mes(request: Request, mes: str = Form(...), data_limite: str = Form(...)):
    """
    Coordenador escolhe no calendário a data-limite para o mês selecionado
    (ex: mês de referência = 06/2026, data_limite = 23/07/2026).
    Essa data específica passa a valer no lugar do dia-limite padrão.
    """
    coordenador = supervisor_logado(request)
    if not coordenador:
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    mes_ref = resolver_mes_escolhido(mes)
    try:
        data_limite_convertida = date.fromisoformat(data_limite)
    except ValueError:
        raise HTTPException(status_code=400, detail="Data inválida.")

    repositorio.definir_prazo_do_mes(mes_ref, data_limite_convertida, coordenador)
    if data_limite_convertida >= date.today():
        repositorio.marcar_solicitacoes_atendidas_do_mes(mes_ref, coordenador)

    return RedirectResponse(f"/coordenador?mes={mes_ref.strftime('%Y-%m')}", status_code=303)


@app.post("/coordenador/{supervisor}/autorizar")
def autorizar_avaliacao_atrasada(request: Request, supervisor: str, mes: str = Form(...)):
    coordenador = supervisor_logado(request)
    if not coordenador:
        return RedirectResponse("/login", status_code=303)
    if request.session.get("tipo") != "coordenador":
        raise HTTPException(status_code=403, detail="Acesso restrito ao coordenador geral.")

    mes_ref = resolver_mes_escolhido(mes)
    repositorio.autorizar_avaliacao_atrasada(supervisor, mes_ref, coordenador)
    repositorio.marcar_solicitacoes_atendidas(supervisor, mes_ref, coordenador)

    return RedirectResponse(f"/coordenador/{supervisor}?mes={mes_ref.strftime('%Y-%m')}", status_code=303)


# ══════════════════════════════════════════════════════════
# FORMULÁRIO DE AVALIAÇÃO
# ══════════════════════════════════════════════════════════
def resolver_mes_escolhido(mes: str | None) -> date:
    """
    Converte o parâmetro `mes` (ex: "2026-05" ou "2026-05-01") vindo da URL/form
    em uma data válida, MAS só aceita meses da lista de meses permitidos
    (mês corrente ou anterior, dentro da janela liberada) — nunca um valor
    arbitrário digitado na URL. Se não vier nada, ou vier algo inválido,
    cai no padrão (mês anterior ao atual).
    """
    permitidos = repositorio.meses_disponiveis_para_avaliacao()
    if not mes:
        return permitidos[0]
    try:
        partes = mes.split("-")
        candidato = date(int(partes[0]), int(partes[1]), 1)
    except (ValueError, IndexError):
        return permitidos[0]
    return candidato if candidato in permitidos else permitidos[0]


@app.post("/solicitar-prazo/{tecnico}")
def solicitar_prazo(request: Request, tecnico: str, mes: str = Form(...)):
    """Supervisor pede ao coordenador para reabrir o prazo de um mês já encerrado."""
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)

    mes_ref = resolver_mes_escolhido(mes)
    repositorio.criar_solicitacao_prazo(supervisor, tecnico, mes_ref)

    return RedirectResponse(f"/avaliar/{tecnico}?mes={mes_ref.strftime('%Y-%m')}", status_code=303)


@app.get("/avaliar/{tecnico}")
def tela_avaliar(request: Request, tecnico: str, mes: str | None = Query(default=None)):
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)

    mes_ref = resolver_mes_escolhido(mes)
    meses_disponiveis = lista_meses_para_template()

    tecnicos_permitidos = repositorio.tecnicos_do_supervisor_no_mes(supervisor, mes_ref)
    tecnico_real = repositorio.encontrar_tecnico(tecnico, tecnicos_permitidos)
    if tecnico_real is None:
        raise HTTPException(status_code=403, detail="Você não pode avaliar este técnico.")
    tecnico = tecnico_real  # a partir daqui, sempre o nome exato como está no banco

    if avaliacoes.ja_avaliado(supervisor, tecnico, mes_ref):
        return templates.TemplateResponse(
            "ja_avaliado.html",
            {
                "request": request,
                "tecnico": tecnico,
                "mes_referencia": mes_ref.strftime("%m/%Y"),
                "mes_referencia_valor": mes_ref.strftime("%Y-%m"),
                "meses_disponiveis": meses_disponiveis,
            },
        )

    if not repositorio.mes_liberado_para_supervisor(supervisor, mes_ref):
        return templates.TemplateResponse(
            "prazo_encerrado.html",
            {
                "request": request,
                "tecnico": tecnico,
                "mes_referencia": mes_ref.strftime("%m/%Y"),
                "mes_referencia_valor": mes_ref.strftime("%Y-%m"),
                "data_limite": repositorio.data_limite_do_mes(mes_ref).strftime("%d/%m/%Y"),
                "meses_disponiveis": meses_disponiveis,
                "solicitacao_pendente": repositorio.existe_solicitacao_pendente(supervisor, mes_ref),
            },
        )

    resumo_visitas = repositorio.resumo_visitas_tecnico(tecnico, mes_ref)

    return templates.TemplateResponse(
        "formulario.html",
        {
            "request": request,
            "supervisor": supervisor,
            "tecnico": tecnico,
            "mes_referencia": mes_ref.strftime("%m/%Y"),
            "mes_referencia_valor": mes_ref.strftime("%Y-%m"),
            "meses_disponiveis": meses_disponiveis,
            "resumo_visitas": resumo_visitas,
        },
    )


@app.post("/avaliar/{tecnico}")
async def salvar_avaliacao(request: Request, tecnico: str):
    supervisor = supervisor_logado(request)
    if not supervisor:
        return RedirectResponse("/login", status_code=303)

    form = await request.form()
    mes_ref = resolver_mes_escolhido(form.get("mes_referencia"))

    tecnicos_permitidos = repositorio.tecnicos_do_supervisor_no_mes(supervisor, mes_ref)
    tecnico_real = repositorio.encontrar_tecnico(tecnico, tecnicos_permitidos)
    if tecnico_real is None:
        raise HTTPException(status_code=403, detail="Você não pode avaliar este técnico.")
    tecnico = tecnico_real  # a partir daqui, sempre o nome exato como está no banco

    if avaliacoes.ja_avaliado(supervisor, tecnico, mes_ref):
        return RedirectResponse("/avaliacoes", status_code=303)

    if not repositorio.mes_liberado_para_supervisor(supervisor, mes_ref):
        # Prazo encerrado e sem autorização do coordenador: barra o envio
        # mesmo que a pessoa tenha deixado a tela do formulário aberta
        # desde antes do prazo acabar.
        raise HTTPException(
            status_code=403,
            detail="O prazo para avaliar este mês encerrou. Peça autorização ao coordenador.",
        )

    respostas = {campo: form.get(campo) for campo in avaliacoes.TODOS_OS_CAMPOS}

    media = avaliacoes.salvar_avaliacao(supervisor, tecnico, mes_ref, respostas)

    return templates.TemplateResponse(
        "sucesso.html",
        {
            "request": request, "tecnico": tecnico, "media": media,
            "classificacao": avaliacoes.classificar(media),
            "mes_referencia_valor": mes_ref.strftime("%Y-%m"),
        },
    )


# ══════════════════════════════════════════════════════════
# COMPARAR AVALIAÇÕES DE UM TÉCNICO ENTRE MESES
# ══════════════════════════════════════════════════════════
@app.get("/comparar/{tecnico}")
def comparar_avaliacoes(request: Request, tecnico: str, sup: str | None = Query(default=None)):
    supervisor_logado_nome = supervisor_logado(request)
    if not supervisor_logado_nome:
        return RedirectResponse("/login", status_code=303)

    eh_coordenador = request.session.get("tipo") == "coordenador"
    if eh_coordenador:
        # Coordenador geral pode ver qualquer técnico, de qualquer supervisor,
        # desde que informe de qual supervisor é (vem do link do histórico geral).
        if not sup:
            raise HTTPException(status_code=400, detail="Supervisor não informado.")
        supervisor = sup
        tecnico_real = tecnico
    else:
        supervisor = supervisor_logado_nome
        tecnicos_permitidos = repositorio.listar_tecnicos_avaliados_pelo_supervisor(supervisor)
        tecnico_real = repositorio.encontrar_tecnico(tecnico, tecnicos_permitidos)
        if tecnico_real is None:
            raise HTTPException(status_code=403, detail="Você não pode ver este técnico.")

    lista_avaliacoes = repositorio.avaliacoes_do_tecnico(supervisor, tecnico_real)
    if not lista_avaliacoes:
        return templates.TemplateResponse(
            "comparacao.html",
            {
                "request": request, "tecnico": tecnico_real,
                "meses": [], "itens": [], "media_valores": [],
                "labels_grafico": "[]", "valores_grafico": "[]", "crescimento": None,
            },
        )

    meses, itens, media_valores = avaliacoes.montar_tabela_resumida(lista_avaliacoes)

    # Dados para o gráfico de evolução (linha do tempo da média final por mês).
    valores_numericos = [float(v) if v not in (None, "") else None for v in media_valores]
    labels_grafico = json.dumps(meses)
    valores_grafico = json.dumps(valores_numericos)

    crescimento = None
    validos = [v for v in valores_numericos if v is not None]
    if len(validos) >= 2:
        crescimento = round(validos[-1] - validos[0], 2)

    return templates.TemplateResponse(
        "comparacao.html",
        {
            "request": request,
            "tecnico": tecnico_real,
            "meses": meses,
            "itens": itens,
            "media_valores": media_valores,
            "labels_grafico": labels_grafico,
            "valores_grafico": valores_grafico,
            "crescimento": crescimento,
        },
    )