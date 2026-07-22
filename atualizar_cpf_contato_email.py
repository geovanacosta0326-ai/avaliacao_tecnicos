"""
Atualiza cpf, contato e email na tabela mestra "tecnicos", a partir da
planilha ampla (colunas: Regional, Perfil, CPF Usuario, Usuario,
Nascimento, E-mail, Celular, Telefone, ID Usuario, Ativo).

So ATUALIZA tecnico que ja existe na tabela (por id_tecnico_responsavel)
-- se o id nao existir, nao grava nada (nao cria registro novo). Nao
sobrescreve com vazio: se a planilha nao tiver um dado, mantem o que ja
esta no banco. NAO mexe no campo "ativo" -- isso continua manual,
controlado pela tela do coordenador.

Precisa que a coluna "email" ja exista na tabela "tecnicos". Se ainda
nao existir, o proprio script cria ela (ALTER TABLE ... IF NOT EXISTS,
seguro, nao apaga nada).

USO (rodar na raiz do projeto, onde tem o .env):
    python atualizar_cpf_contato_email.py "tabela.xlsx"
"""
import sys
import openpyxl
from sqlalchemy import text
from app.database import get_engine

COL_CPF = 3
COL_NOME = 4
COL_EMAIL = 6
COL_CELULAR = 7
COL_TELEFONE = 8
COL_ID = 9


def limpa(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def atualizar(caminho: str):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.worksheets[0]

    engine = get_engine()

    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE tecnicos ADD COLUMN IF NOT EXISTS email TEXT;"))

    atualizados = 0
    nao_encontrados = []
    sem_id = 0

    with engine.begin() as conn:
        for r in range(2, ws.max_row + 1):
            id_valor = ws.cell(row=r, column=COL_ID).value
            if id_valor is None:
                sem_id += 1
                continue
            try:
                id_tecnico = int(str(id_valor).strip())
            except ValueError:
                sem_id += 1
                continue

            nome = limpa(ws.cell(row=r, column=COL_NOME).value)
            cpf = limpa(ws.cell(row=r, column=COL_CPF).value)
            email = limpa(ws.cell(row=r, column=COL_EMAIL).value)
            celular = limpa(ws.cell(row=r, column=COL_CELULAR).value)
            telefone = limpa(ws.cell(row=r, column=COL_TELEFONE).value)
            contato = celular or telefone

            resultado = conn.execute(
                text("""
                    UPDATE tecnicos
                    SET cpf = COALESCE(:cpf, cpf),
                        contato = COALESCE(:contato, contato),
                        email = COALESCE(:email, email),
                        atualizado_em = NOW()
                    WHERE id_tecnico_responsavel = :id;
                """),
                {"id": id_tecnico, "cpf": cpf, "contato": contato, "email": email},
            )

            if resultado.rowcount == 0:
                nao_encontrados.append((id_tecnico, nome))
            else:
                atualizados += 1

    print(f"\nAtualizados: {atualizados} tecnico(s).")
    if sem_id:
        print(f"Linhas sem ID valido, ignoradas: {sem_id}")
    if nao_encontrados:
        print(f"\nNAO encontrados na tabela 'tecnicos' (nao existiam, entao nao foram gravados):")
        for id_tecnico, nome in nao_encontrados:
            print(f"   {id_tecnico}  {nome or ''}")
    else:
        print("Todos os IDs da planilha foram encontrados na tabela.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python atualizar_cpf_contato_email.py <planilha.xlsx>")
        sys.exit(1)
    atualizar(sys.argv[1])
