"""
Importa uma planilha "Cadastro dos Tecnicos" (formato de 31 colunas, o
mesmo exportado pelo sistema) direto para a tabela "tecnicos" do banco.

So ATUALIZA tecnico que ja existe (por id_tecnico_responsavel) -- se o id
nao existir na tabela, aquele UPDATE simplesmente nao afeta nenhuma linha
(nao cria registro novo). Nao sobrescreve com vazio: se um campo vier em
branco na planilha, mantem o que ja estiver gravado no banco. Nunca mexe
em nome / ativo / primeira_visita / ultima_visita (isso vem so da
sincronizacao com a base de visitas).

USO (rodar na raiz do projeto, onde tem o .env):
    python importar_cadastro_tecnicos.py "Cadastro dos Tecnicos.xlsx"

Opcionalmente, se voce tiver uma segunda planilha com projeto/atividade/
datas de inicio e fim de vinculo (colunas: Id Tecnico, Atividade/Cadeia,
Projeto, Inicio, Fim), pode passar como segundo argumento e ele cruza
pelos dois:
    python importar_cadastro_tecnicos.py "Cadastro.xlsx" "Atividades.xlsx"

No fim, imprime um resumo: quantos tecnicos foram atualizados, quantos
IDs da planilha nao foram encontrados no banco (esses ficam de fora, por
seguranca -- nunca cria linha nova sozinho).
"""
import sys
import openpyxl
from sqlalchemy import text
from app.database import get_engine

# Colunas esperadas na planilha de cadastro (31 colunas, no padrao que o
# sistema ja exporta). Se o layout mudar, ajuste os indices aqui.
COL_ID = 1
COL_NOME = 3
COL_CPF = 4
COL_CELULAR = 6
COL_TELEFONE = 7
COL_RG = 9
COL_LOGRADOURO = 12
COL_NUMERO = 13
COL_COMPLEMENTO = 14
COL_BAIRRO = 16
COL_MUNICIPIO = 17
COL_NOME_FANTASIA = 21
COL_CNPJ = 22


def carregar_atividades(caminho: str | None) -> dict:
    """Le uma planilha auxiliar de atividade/projeto/datas, indexada por
    Id Tecnico. Formato esperado: Id Tecnico, Atividade (ou Cadeia),
    Projeto, Inicio, Fim -- em qualquer ordem, desde que o cabecalho bata
    com esses nomes (nao sensivel a acento/maiusculas)."""
    if not caminho:
        return {}

    def normaliza(s):
        import unicodedata
        s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
        return s.strip().lower()

    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.worksheets[0]
    headers = [normaliza(c.value or "") for c in ws[1]]

    def achar(*nomes):
        for nome in nomes:
            if nome in headers:
                return headers.index(nome)
        return None

    idx_id = achar("id tecnico", "id_tecnico")
    idx_ativ = achar("atividade", "cadeia")
    idx_proj = achar("projeto")
    idx_inicio = achar("inicio", "data inicio", "data_visita_inicial")
    idx_fim = achar("fim", "data fim", "data_visita_final")

    resultado = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if idx_id is None or row[idx_id] is None:
            continue
        chave = str(row[idx_id]).strip()
        resultado[chave] = {
            "atividade": row[idx_ativ] if idx_ativ is not None else None,
            "projeto": row[idx_proj] if idx_proj is not None else None,
            "inicio": row[idx_inicio] if idx_inicio is not None else None,
            "fim": row[idx_fim] if idx_fim is not None else None,
        }
    return resultado


def montar_endereco(logradouro, numero, complemento, bairro) -> str | None:
    partes = [str(p).strip() for p in [logradouro, numero, complemento, bairro] if p]
    return ", ".join(partes) if partes else None


def limpa(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def importar(caminho_cadastro: str, caminho_atividades: str | None = None):
    atividades = carregar_atividades(caminho_atividades)

    wb = openpyxl.load_workbook(caminho_cadastro, data_only=True)
    ws = wb.worksheets[0]

    engine = get_engine()
    atualizados = 0
    nao_encontrados = []

    with engine.begin() as conn:
        for r in range(2, ws.max_row + 1):
            id_valor = ws.cell(row=r, column=COL_ID).value
            if id_valor is None or str(id_valor).strip().upper().startswith("TOTAL"):
                continue

            chave = str(id_valor).strip()
            try:
                id_tecnico = int(chave)
            except ValueError:
                continue

            nome = limpa(ws.cell(row=r, column=COL_NOME).value)
            cpf = limpa(ws.cell(row=r, column=COL_CPF).value)
            celular = limpa(ws.cell(row=r, column=COL_CELULAR).value)
            telefone = limpa(ws.cell(row=r, column=COL_TELEFONE).value)
            rg = limpa(ws.cell(row=r, column=COL_RG).value)
            logradouro = ws.cell(row=r, column=COL_LOGRADOURO).value
            numero = ws.cell(row=r, column=COL_NUMERO).value
            complemento = ws.cell(row=r, column=COL_COMPLEMENTO).value
            bairro = ws.cell(row=r, column=COL_BAIRRO).value
            municipio = limpa(ws.cell(row=r, column=COL_MUNICIPIO).value)
            nome_fantasia = limpa(ws.cell(row=r, column=COL_NOME_FANTASIA).value)
            cnpj = limpa(ws.cell(row=r, column=COL_CNPJ).value)

            contato = celular or telefone
            endereco = montar_endereco(logradouro, numero, complemento, bairro)

            ativ = atividades.get(chave, {})
            inicio = limpa(ativ.get("inicio"))
            fim = limpa(ativ.get("fim"))

            resultado = conn.execute(
                text("""
                    UPDATE tecnicos
                    SET rg = COALESCE(:rg, rg),
                        cpf = COALESCE(:cpf, cpf),
                        contato = COALESCE(:contato, contato),
                        empresa = COALESCE(:empresa, empresa),
                        cnpj_empresa = COALESCE(:cnpj_empresa, cnpj_empresa),
                        endereco = COALESCE(:endereco, endereco),
                        municipio = COALESCE(:municipio, municipio),
                        data_inicio_vinculo = COALESCE(:inicio, data_inicio_vinculo),
                        data_fim_vinculo = COALESCE(:fim, data_fim_vinculo),
                        atualizado_em = NOW()
                    WHERE id_tecnico_responsavel = :id;
                """),
                {
                    "id": id_tecnico,
                    "rg": rg,
                    "cpf": cpf,
                    "contato": contato,
                    "empresa": nome_fantasia,
                    "cnpj_empresa": cnpj,
                    "endereco": endereco,
                    "municipio": municipio,
                    "inicio": inicio,
                    "fim": fim,
                },
            )

            if resultado.rowcount == 0:
                nao_encontrados.append((id_tecnico, nome))
            else:
                atualizados += 1

    print(f"\nAtualizados: {atualizados} tecnico(s).")
    if nao_encontrados:
        print(f"\nNAO encontrados na tabela 'tecnicos' (nao foram gravados -- id nao existe):")
        for id_tecnico, nome in nao_encontrados:
            print(f"   {id_tecnico}  {nome or ''}")
    else:
        print("Todos os IDs da planilha foram encontrados na tabela.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python importar_cadastro_tecnicos.py <planilha_cadastro.xlsx> [planilha_atividades.xlsx]")
        sys.exit(1)

    caminho_cadastro = sys.argv[1]
    caminho_atividades = sys.argv[2] if len(sys.argv) > 2 else None
    importar(caminho_cadastro, caminho_atividades)
