"""
Importa as planilhas de supervisores e tecnicos para as tabelas mestras
(supervisores e tecnicos). Roda quantas vezes precisar - atualiza quem
ja existe (por nome/id) sem duplicar, e nao mexe no campo "ativo" de
quem ja estiver cadastrado (so entra como ativo=True na primeira vez).

Rode assim, na raiz do projeto (onde tem o .env):
    python importar_supervisores_tecnicos.py caminho/da/planilha.xlsx

A planilha deve ter duas abas:
  - Uma aba com uma coluna "supervisor" (lista de nomes)
  - Uma aba com as colunas "id_tecnico_responsavel", "tecnico_responsavel",
    "primeira_visita", "ultima_visita"

O script detecta sozinho qual aba é qual, pelo cabeçalho.
"""
import sys
import openpyxl
from sqlalchemy import text
from app.database import get_engine

if len(sys.argv) != 2:
    print("Uso: python importar_supervisores_tecnicos.py caminho/da/planilha.xlsx")
    sys.exit(1)

caminho = sys.argv[1]
wb = openpyxl.load_workbook(caminho, data_only=True)

aba_supervisores = None
aba_tecnicos = None

for nome_aba in wb.sheetnames:
    ws = wb[nome_aba]
    primeira_linha = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not primeira_linha or primeira_linha[0] is None:
        continue
    cabecalho = [str(c).strip().lower() if c else "" for c in primeira_linha]
    if cabecalho == ["supervisor"]:
        aba_supervisores = ws
    elif "id_tecnico_responsavel" in cabecalho:
        aba_tecnicos = ws

if aba_supervisores is None:
    print("Não achei a aba de supervisores (esperava uma coluna 'supervisor').")
if aba_tecnicos is None:
    print("Não achei a aba de técnicos (esperava a coluna 'id_tecnico_responsavel').")
if aba_supervisores is None or aba_tecnicos is None:
    sys.exit(1)

engine = get_engine()

# ── Supervisores ──────────────────────────────────────────
qtd_supervisores = 0
with engine.begin() as conn:
    for row in aba_supervisores.iter_rows(min_row=2, values_only=True):
        nome = (row[0] or "").strip() if row and row[0] else None
        if not nome:
            continue
        conn.execute(
            text("""
                INSERT INTO supervisores (nome)
                VALUES (:nome)
                ON CONFLICT (nome) DO NOTHING;
            """),
            {"nome": nome},
        )
        qtd_supervisores += 1

print(f"Supervisores processados: {qtd_supervisores}")

# ── Técnicos ──────────────────────────────────────────────
qtd_tecnicos = 0
with engine.begin() as conn:
    for row in aba_tecnicos.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        id_tecnico, nome, primeira_visita, ultima_visita = row[0], row[1], row[2], row[3]
        nome = (nome or "").strip()
        if not nome:
            continue
        conn.execute(
            text("""
                INSERT INTO tecnicos (id_tecnico_responsavel, nome, primeira_visita, ultima_visita)
                VALUES (:id, :nome, :primeira_visita, :ultima_visita)
                ON CONFLICT (id_tecnico_responsavel) DO UPDATE
                SET nome = EXCLUDED.nome,
                    primeira_visita = EXCLUDED.primeira_visita,
                    ultima_visita = EXCLUDED.ultima_visita,
                    atualizado_em = NOW();
            """),
            {
                "id": int(id_tecnico),
                "nome": nome,
                "primeira_visita": primeira_visita,
                "ultima_visita": ultima_visita,
            },
        )
        qtd_tecnicos += 1

print(f"Técnicos processados: {qtd_tecnicos}")
print("\nImportação concluída.")
